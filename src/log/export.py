import sqlite3
import logging
import pandas as pd
from typing import Any
from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS

log = logging.getLogger(__name__)

# Define styles as constants for reuse
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFA0", end_color="FFFFA0", fill_type="solid")
RED_FILL = PatternFill(start_color="FF9696", end_color="FF9696", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E6E6FF", end_color="E6E6FF", fill_type="solid")
TEXT_FORMAT = BUILTIN_FORMATS[49]  # '@' (Text format)


def escape_formula(value: Any) -> Any:
    """
    Prepends a single quote to a string if it starts with a character
    that Excel might interpret as a formula, to prevent formula injection.

    :param value: The value to check and potentially escape.
    :return: The escaped string or the original value if no escape was needed.
    """
    if isinstance(value, str) and value.startswith(('=', '-', '+', '@')):
        return f"'{value}"
    return value

def get_logs_from_database(db_path: Path) -> pd.DataFrame:
    """
    Retrieve log data from the SQLite database.
    
    :param db_path: The file path to the SQLite database.
    :return: DataFrame with log entries or None if an error occurred.
    """
    log.debug(f"Connecting to database: {db_path}")
    try:
        with sqlite3.connect(db_path) as con:
            query = """
            SELECT 
                datetime(timestamp, 'unixepoch', 'localtime') as Timestamp,
                level as 'Log Level',
                module as 'Module',
                funcName || ':' || lineno as 'Func Source',
                message as 'Message'
            FROM logs
            ORDER BY timestamp ASC;
            """
            df = pd.read_sql_query(query, con)
            log.info(f"Read {len(df)} log entries from the database.")
            return df
    except (sqlite3.Error, pd.errors.DatabaseError) as e:
        log.error(f"An error occurred while reading the database: {e}")
        return None

def sanitize_log_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sanitize data to prevent Excel formula injection.
    
    :param df: DataFrame with log entries.
    :return: Sanitized DataFrame.
    """
    log.debug("Sanitizing data to prevent Excel formula errors...")
    for col in ['Module', 'Func Source', 'Message']:
        df[col] = df[col].apply(escape_formula)
    return df

def style_header(ws):
    """
    Apply styling to the header row.
    
    :param ws: Excel worksheet.
    """
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

def apply_row_styling(ws, level_col_idx: int):
    """
    Apply conditional formatting based on log level.
    
    :param ws: Excel worksheet.
    :param level_col_idx: Index of the log level column.
    """
    fill_map = {
        'WARNING': YELLOW_FILL,
        'ERROR': RED_FILL,
        'CRITICAL': RED_FILL,
        'FATAL': RED_FILL,
        'DEBUG': BLUE_FILL
    }
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        log_level = row[level_col_idx - 1].value
        fill_to_apply = fill_map.get(log_level)
        if fill_to_apply:
            for cell in row:
                cell.fill = fill_to_apply
        
        # Apply text format to all but the timestamp column
        for cell in row[1:]:
            cell.number_format = TEXT_FORMAT

def adjust_column_widths(ws):
    """
    Auto-adjust column widths based on content.
    
    :param ws: Excel worksheet.
    """
    column_widths = {}
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            if cell.value:
                column_widths[i] = max(column_widths.get(i, 0), len(str(cell.value)))
    
    for i, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = width + 2

def write_to_excel(df: pd.DataFrame, output_path: Path) -> bool:
    """
    Write data to Excel with styling.
    
    :param df: DataFrame with log entries.
    :param output_path: Path where Excel file will be saved.
    :return: True if successful, False otherwise.
    """
    log.info(f"Writing data to Excel file: {output_path}")
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Logs')
            ws = writer.sheets['Logs']

            # Apply styling
            style_header(ws)
            level_col_idx = df.columns.get_loc('Log Level') + 1
            apply_row_styling(ws, level_col_idx)
            adjust_column_widths(ws)

        log.info("\nExport successful!")
        log.info(f"File saved to: {output_path.resolve()}")
        return True
    except Exception as e:
        log.error(f"An error occurred while writing or styling the Excel file: {e}")
        return False

def export_logs_to_excel(db_path: Path, output_path: Path) -> None:
    """
    Exports log entries from the SQLite database to a styled Excel file.

    Features include auto-sized columns, conditional row coloring based on log level,
    and protection against Excel formula injection.

    :param db_path: The file path to the SQLite database.
    :param output_path: The file path where the Excel file will be saved.
    """
    if not db_path.exists():
        log.error(f"Error: Database file not found at '{db_path}'")
        return

    # Get data from database
    df = get_logs_from_database(db_path)
    if df is None:
        return
        
    if df.empty:
        log.warning("No log entries to export.")
        return

    # Sanitize data
    df = sanitize_log_data(df)

    # Write to Excel with styling
    write_to_excel(df, output_path)
